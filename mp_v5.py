import os
import time
import threading
from concurrent.futures import ThreadPoolExecutor, as_completed
from flask import Flask, request, render_template_string, send_file, flash, redirect, url_for, jsonify
import pandas as pd
import requests
from openpyxl import load_workbook
from openpyxl.styles import Border, Side, Font
from datetime import datetime

app = Flask(__name__)
app.secret_key = 'your_secret_key'

API_TOKEN = "f9be4985f5e3488716c040ca52b8e04c7c0f9e0b".strip()
API_TOKEN = API_TOKEN.encode('ascii', errors='ignore').decode()
STORE_ID = "241ed919-a631-11ee-0a80-07a9000bb947"

# Константы для создания заказа
ORGANIZATION_UUID = "4bf22d14-4d5e-11ee-0a80-0761000a555b"
COUNTERPARTY_UUID = "5ba713c4-a31d-11ee-0a80-063f0084f98f"
STORE_UUID = "241ed919-a631-11ee-0a80-07a9000bb947"
PROJECT_UUID = "4ec39020-4e1d-11ee-0a80-00c60006dca7"

HEADERS = {
    "Authorization": f"Bearer {API_TOKEN}",
    "Accept-Encoding": "gzip",
    "Content-Type": "application/json"
}

UPLOAD_FOLDER = 'uploads'
RESULT_FOLDER = 'results'
os.makedirs(UPLOAD_FOLDER, exist_ok=True)
os.makedirs(RESULT_FOLDER, exist_ok=True)

progress = {}
cancel_flags = {}
order_progress = {}  # Отдельный прогресс для создания заказов

HEADER_HTML = '''
<div style="padding:10px; background:#f0f0f0; border-bottom:1px solid #ccc; display:flex; justify-content:space-between; align-items:center;">
  <div style="display:flex; align-items:center;">
    <button onclick="goHome()" style="margin-right:20px;">Главная</button>
    <strong>Формирование отчета для сборки товаров с учетом ячеек</strong>
  </div>
  <div>
    <!-- Дополнительные кнопки справа при необходимости -->
  </div>
</div>
<script>
function goHome() {
  const sid = window.sessionStorage.getItem('currentSession');
  if (sid) {
    fetch(`/status/${sid}`)
      .then(res => res.json())
      .then(data => {
        // Проверяем, действительно ли процесс активен (не завершен и не в состоянии ошибки)
        const status = data.status.toLowerCase();
        const isActive = !status.includes('завершена') && 
                        !status.includes('ошибка') && 
                        !status.includes('отменена') && 
                        !status.includes('нет данных') &&
                        status.includes('обрабатываем') || 
                        status.includes('получаем') || 
                        status.includes('начинаем') ||
                        status.includes('формируем');
        
        if (isActive) {
          alert('Дождитесь окончания текущей обработки или нажмите "Остановить процесс"');
        } else {
          // Очищаем sessionStorage если процесс завершен
          window.sessionStorage.removeItem('currentSession');
          window.location.href = '/';
        }
      })
      .catch(error => {
        // Если ошибка запроса (например, сессия не найдена), просто переходим на главную
        window.sessionStorage.removeItem('currentSession');
        window.location.href = '/';
      });
  } else {
    window.location.href = '/';
  }
}
</script>
'''

# --- Новый функционал: ограничение количества файлов ---
def clean_old_results(max_files=50):
    """Удаляет старые файлы в папке results, если их больше max_files."""
    files = [os.path.join(RESULT_FOLDER, f) for f in os.listdir(RESULT_FOLDER) if f.endswith('.xlsx')]
    files.sort(key=lambda x: os.path.getmtime(x), reverse=True)
    for f in files[max_files:]:
        try:
            os.remove(f)
        except Exception:
            pass

# --- Функции для создания заказа ---
def get_product_uuid_for_order(article):
    """Получает UUID товара по артикулу для создания заказа."""
    url = "https://api.moysklad.ru/api/remap/1.2/entity/product"
    params = {"filter": f"article={article}", "limit": 1}
    resp = requests.get(url, headers=HEADERS, params=params)
    resp.raise_for_status()
    data = resp.json()
    rows = data.get('rows', [])
    if not rows:
        return None
    return rows[0]['id']

def create_customer_order_from_file(filepath, session_id):
    """Создает заказ покупателя на основе данных из Excel файла с отладкой."""
    try:
        order_progress[session_id] = "🔄 Начинаем создание заказа..."
        print(f"[ORDER {session_id}] Начинаем создание заказа из файла: {filepath}", flush=True)
        
        if cancel_flags.get(f"order_{session_id}"):
            order_progress[session_id] = "❌ Создание заказа отменено пользователем"
            return {"error": "Создание заказа отменено пользователем"}
        
        # Читаем Excel файл
        order_progress[session_id] = "📖 Читаем Excel файл..."
        print(f"[ORDER {session_id}] Читаем файл...", flush=True)
        df = pd.read_excel(filepath)
        print(f"[ORDER {session_id}] Файл прочитан, строк: {len(df)}", flush=True)
        
        if cancel_flags.get(f"order_{session_id}"):
            order_progress[session_id] = "❌ Создание заказа отменено пользователем"
            return {"error": "Создание заказа отменено пользователем"}
        
        # Проверяем наличие необходимых колонок
        order_progress[session_id] = "🔍 Проверяем колонки файла..."
        print(f"[ORDER {session_id}] Колонки в файле: {list(df.columns)}", flush=True)
        required_columns = ['Артикул', 'Количество']
        for col in required_columns:
            if col not in df.columns:
                error_msg = f"Не найдена колонка '{col}' в файле"
                order_progress[session_id] = f"❌ {error_msg}"
                print(f"[ORDER {session_id}] ОШИБКА: {error_msg}", flush=True)
                return {"error": error_msg}
        
        # Фильтруем строки с валидными данными
        order_progress[session_id] = "📋 Фильтруем валидные товары..."
        valid_rows = []
        for idx, row in df.iterrows():
            if cancel_flags.get(f"order_{session_id}"):
                order_progress[session_id] = "❌ Создание заказа отменено пользователем"
                return {"error": "Создание заказа отменено пользователем"}
                
            article = str(row['Артикул']).strip()
            quantity = row['Количество']
            
            if article and article != 'nan' and pd.notna(quantity) and quantity > 0:
                valid_rows.append({
                    'article': article,
                    'quantity': int(quantity)
                })
                print(f"[ORDER {session_id}] Валидная строка {idx+1}: {article} x {int(quantity)}", flush=True)
        
        print(f"[ORDER {session_id}] Найдено валидных товаров: {len(valid_rows)}", flush=True)
        if not valid_rows:
            error_msg = "Не найдено товаров для добавления в заказ"
            order_progress[session_id] = f"❌ {error_msg}"
            return {"error": error_msg}
        
        # Получаем UUID для всех товаров
        order_progress[session_id] = f"🔍 Ищем товары в МойСклад... (0/{len(valid_rows)})"
        positions = []
        not_found_articles = []
        
        for i, item in enumerate(valid_rows):
            if cancel_flags.get(f"order_{session_id}"):
                order_progress[session_id] = "❌ Создание заказа отменено пользователем"
                return {"error": "Создание заказа отменено пользователем"}
                
            order_progress[session_id] = f"🔍 Ищем товары в МойСклад... ({i+1}/{len(valid_rows)}) - {item['article']}"
            print(f"[ORDER {session_id}] Ищем товар {i+1}/{len(valid_rows)}: {item['article']}", flush=True)
            
            try:
                product_uuid = get_product_uuid_for_order(item['article'])
                if product_uuid:
                    positions.append({
                        "assortment": {
                            "meta": {
                                "href": f"https://api.moysklad.ru/api/remap/1.2/entity/product/{product_uuid}",
                                "type": "product"
                            }
                        },
                        "quantity": item['quantity'],
                        "price": 0,
                        "vat": 20,
                        "vatEnabled": True,
                        "discount": 0,
                        "reserve": 0
                    })
                    print(f"[ORDER {session_id}] ✅ Найден: {item['article']} -> {product_uuid}", flush=True)
                else:
                    not_found_articles.append(item['article'])
                    print(f"[ORDER {session_id}] ❌ НЕ найден: {item['article']}", flush=True)
            except Exception as e:
                not_found_articles.append(item['article'])
                print(f"[ORDER {session_id}] ❌ Ошибка поиска {item['article']}: {e}", flush=True)
            
            time.sleep(0.1)  # Небольшая задержка для возможности отмены
        
        if cancel_flags.get(f"order_{session_id}"):
            order_progress[session_id] = "❌ Создание заказа отменено пользователем"
            return {"error": "Создание заказа отменено пользователем"}
        
        if not positions:
            error_msg = f"Ни один из артикулов не найден в системе: {', '.join(not_found_articles)}"
            order_progress[session_id] = f"❌ {error_msg}"
            print(f"[ORDER {session_id}] ОШИБКА: {error_msg}", flush=True)
            return {"error": error_msg}
        
        print(f"[ORDER {session_id}] Найдено товаров: {len(positions)}, не найдено: {len(not_found_articles)}", flush=True)
        
        # Создаем заказ
        order_progress[session_id] = "📝 Создаем заказ в МойСклад..."
        url = "https://api.moysklad.ru/api/remap/1.2/entity/customerorder"
        now_iso = datetime.utcnow().strftime("%Y-%m-%d %H:%M:%S")
        
        order_body = {
            "name": f"Автозаказ {datetime.now().strftime('%d.%m.%Y %H:%M')}",
            "moment": now_iso,
            "organization": {
                "meta": {
                    "href": f"https://api.moysklad.ru/api/remap/1.2/entity/organization/{ORGANIZATION_UUID}",
                    "type": "organization"
                }
            },
            "agent": {
                "meta": {
                    "href": f"https://api.moysklad.ru/api/remap/1.2/entity/counterparty/{COUNTERPARTY_UUID}",
                    "type": "counterparty"
                }
            },
            "store": {
                "meta": {
                    "href": f"https://api.moysklad.ru/api/remap/1.2/entity/store/{STORE_UUID}",
                    "type": "store"
                }
            },
            "project": {
                "meta": {
                    "href": f"https://api.moysklad.ru/api/remap/1.2/entity/project/{PROJECT_UUID}",
                    "type": "project"
                }
            },
            "currency": {
                "meta": {
                    "href": "https://api.moysklad.ru/api/remap/1.2/entity/currency/643",
                    "type": "currency"
                }
            },
            "positions": positions,
            "description": f"Создан автоматически из файла. Добавлено позиций: {len(positions)}" + 
                         (f". Не найдены артикулы: {', '.join(not_found_articles)}" if not_found_articles else "")
        }
        
        print(f"[ORDER {session_id}] Отправляем POST запрос в МойСклад...", flush=True)
        print(f"[ORDER {session_id}] URL: {url}", flush=True)
        print(f"[ORDER {session_id}] Тело запроса: {order_body}", flush=True)
        
        if cancel_flags.get(f"order_{session_id}"):
            order_progress[session_id] = "❌ Создание заказа отменено пользователем"
            return {"error": "Создание заказа отменено пользователем"}
        
        resp = requests.post(url, headers=HEADERS, json=order_body, timeout=30)
        print(f"[ORDER {session_id}] Ответ сервера: статус {resp.status_code}", flush=True)
        print(f"[ORDER {session_id}] Ответ сервера: {resp.text}", flush=True)
        
        resp.raise_for_status()
        order_data = resp.json()
        
        result = {
            "success": True,
            "order_id": order_data.get('id'),
            "order_name": order_data.get('name'),
            "positions_added": len(positions),
            "total_items": len(valid_rows)
        }
        
        if not_found_articles:
            result["not_found_articles"] = not_found_articles
        
        order_progress[session_id] = f"✅ Заказ создан успешно! ID: {result['order_id']}"
        print(f"[ORDER {session_id}] ✅ Заказ создан успешно: {result}", flush=True)
        return result
        
    except requests.exceptions.HTTPError as e:
        error_msg = f"Ошибка API МойСклад: {e.response.status_code} - {e.response.text}"
        order_progress[session_id] = f"❌ {error_msg}"
        print(f"[ORDER {session_id}] HTTP ОШИБКА: {error_msg}", flush=True)
        return {"error": error_msg}
    except requests.exceptions.Timeout:
        error_msg = "Превышено время ожидания ответа от МойСклад (30 сек)"
        order_progress[session_id] = f"❌ {error_msg}"
        print(f"[ORDER {session_id}] TIMEOUT: {error_msg}", flush=True)
        return {"error": error_msg}
    except Exception as e:
        error_msg = f"Ошибка создания заказа: {str(e)}"
        order_progress[session_id] = f"❌ {error_msg}"
        print(f"[ORDER {session_id}] ОБЩАЯ ОШИБКА: {error_msg}", flush=True)
        return {"error": error_msg}

# --- Функция для получения списка последних файлов ---
def get_recent_files(count=10):
    """Возвращает список последних созданных файлов с метаданными."""
    if not os.path.exists(RESULT_FOLDER):
        return []
    
    files = []
    for filename in os.listdir(RESULT_FOLDER):
        if filename.endswith('.xlsx'):
            filepath = os.path.join(RESULT_FOLDER, filename)
            try:
                mtime = os.path.getmtime(filepath)
                size = os.path.getsize(filepath)
                files.append({
                    'filename': filename,
                    'filepath': filepath,
                    'mtime': mtime,
                    'size': size,
                    'formatted_time': time.strftime('%d.%m.%Y %H:%M:%S', time.localtime(mtime)),
                    'formatted_size': f"{size / 1024:.1f} KB" if size < 1024*1024 else f"{size / (1024*1024):.1f} MB"
                })
            except OSError:
                continue
    
    # Сортируем по времени создания (новые сверху)
    files.sort(key=lambda x: x['mtime'], reverse=True)
    return files[:count]

# --- Вспомогательные функции (без изменений) ---
def find_column_index(columns, names):
    for idx, col in enumerate(columns):
        col_lower = str(col).strip().lower()
        for name in names:
            if col_lower == name.lower():
                return idx
    return None

def find_quantity_column(df, candidates=['кол-во', 'количество', 'кол']):
    for col in df.columns[:10]:
        col_lower = str(col).strip().lower()
        if any(name == col_lower for name in candidates):
            sample = df[col].dropna().head(10)
            if not sample.empty and pd.to_numeric(sample, errors='coerce').notnull().any():
                return df.columns.get_loc(col)
    return None

def get_product_uuid(article):
    url = "https://api.moysklad.ru/api/remap/1.2/entity/product"
    params = {"filter": f"article={article}", "limit": 1}
    resp = requests.get(url, headers=HEADERS, params=params)
    resp.raise_for_status()
    data = resp.json()
    rows = data.get('rows', [])
    if not rows:
        return None, None
    row = rows[0]
    return row['id'], row.get('name', '')

def get_store_slots(store_id):
    url = f"https://api.moysklad.ru/api/remap/1.2/entity/store/{store_id}/slots"
    resp = requests.get(url, headers=HEADERS, params={"limit":1000})
    resp.raise_for_status()
    data = resp.json()
    return {row['id']: row['name'] for row in data.get('rows', [])}

def get_stock_by_slot(product_uuid, store_id):
    url = "https://api.moysklad.ru/api/remap/1.2/report/stock/byslot/current"
    params = [
        ('filter', f"assortmentId={product_uuid}"),
        ('filter', f"storeId={store_id}"),
        ('limit','1000')
    ]
    resp = requests.get(url, headers=HEADERS, params=params)
    resp.raise_for_status()
    return resp.json()

def process_article(article, slot_names):
    session_id = threading.current_thread().name
    article = str(article).strip()
    if not article or cancel_flags.get(session_id):
        return None, None, ""
    try:
        uuid, name = get_product_uuid(article)
        if not uuid:
            return None, None, ""
        rows = get_stock_by_slot(uuid, STORE_ID)
        parts = []
        for entry in rows:
            if cancel_flags.get(session_id):
                break
            slot_id = entry.get('slotId')
            qty = entry.get('stock', 0)
            if slot_id and qty > 0:
                parts.append(f"{slot_names.get(slot_id, slot_id)} - {int(qty)} шт")
        time.sleep(0.05)
        return article, name, ", ".join(parts)
    except Exception as e:
        print(f"Ошибка для артикула {article}: {e}", flush=True)
        return None, None, ""

def format_sticker_cell(cell):
    """Форматирует ячейку с номером стикера."""
    try:
        value = str(cell.value) if cell.value else ""
        if len(value) < 4:
            return
        main_part = value[:-4].rstrip()
        last_four = value[-4:]
        cell.value = f"{main_part} {last_four}"
        cell.font = cell.font.copy(bold=True, size=cell.font.size + 1)
    except Exception as e:
        print(f"Ошибка форматирования ячейки: {e}", flush=True)

def save_workbook_with_retries(wb, filename, session_id, retries=5, delay=3):
    """Сохраняет workbook с повторными попытками и отладкой."""
    for attempt in range(1, retries+1):
        try:
            progress[session_id] = f"[{session_id}] Попытка сохранения файла {attempt}/{retries}..."
            print(f"[{session_id}] Попытка {attempt}: сохраняем файл {filename}", flush=True)
            wb.save(filename)
            progress[session_id] = f"[{session_id}] Файл успешно сохранён!"
            print(f"[{session_id}] Файл {filename} успешно сохранён.", flush=True)
            return True
        except PermissionError as e:
            error_msg = f"Файл занят другим процессом, жду {delay} сек..."
            progress[session_id] = f"[{session_id}] {error_msg}"
            print(f"[{session_id}] Попытка {attempt}: {error_msg}", flush=True)
            if attempt < retries:
                time.sleep(delay)
        except Exception as e:
            error_msg = f"Ошибка сохранения файла: {str(e)}"
            progress[session_id] = f"[{session_id}] {error_msg}"
            print(f"[{session_id}] Попытка {attempt}: {error_msg}", flush=True)
            if attempt >= retries:
                return False
            time.sleep(1)
    
    progress[session_id] = f"[{session_id}] Не удалось сохранить файл после {retries} попыток"
    print(f"[{session_id}] Не удалось сохранить файл {filename} после {retries} попыток.", flush=True)
    return False

# --- Основной процесс обработки ---
def process_file(input_path, output_path, session_id):
    try:
        cancel_flags[session_id] = False

        progress[session_id] = f"[{session_id}] Начинаем обработку файла"
        print(f"[{session_id}] Начинаем обработку файла: {input_path}", flush=True)
        
        # Читаем Excel файл
        progress[session_id] = f"[{session_id}] Читаем Excel файл..."
        df = pd.read_excel(input_path)
        progress[session_id] = f"[{session_id}] Excel загружен: {len(df)} строк"
        print(f"[{session_id}] Excel загружен: {len(df)} строк, колонки: {list(df.columns)}", flush=True)

        # Проверяем отмену
        if cancel_flags.get(session_id):
            progress[session_id] = f"[{session_id}] Процесс отменён пользователем"
            return

        # Ищем колонки
        progress[session_id] = f"[{session_id}] Ищем необходимые колонки..."
        article_col = find_column_index(df.columns, ['артикул'])
        sticker_col = find_column_index(df.columns, ['№ стикера','номер стикера','стикер','номер'])
        quantity_col = find_quantity_column(df)
        
        print(f"[{session_id}] Найденные колонки - Артикул: {article_col}, Стикер: {sticker_col}, Количество: {quantity_col}", flush=True)
        
        if article_col is None or sticker_col is None or quantity_col is None or cancel_flags.get(session_id):
            progress[session_id] = f"[{session_id}] Ошибка: не найдены обязательные колонки или процесс отменён"
            print(f"[{session_id}] ОШИБКА: не найдены обязательные колонки", flush=True)
            return

        # Получаем ячейки склада
        progress[session_id] = f"[{session_id}] Получаем ячейки склада..."
        print(f"[{session_id}] Запрашиваем ячейки склада...", flush=True)
        slot_names = get_store_slots(STORE_ID)
        progress[session_id] = f"[{session_id}] Ячеек получено: {len(slot_names)}"
        print(f"[{session_id}] Ячеек получено: {len(slot_names)}", flush=True)
        
        if cancel_flags.get(session_id):
            progress[session_id] = f"[{session_id}] Процесс отменён до обработки статей"
            return

        # Обрабатываем артикулы
        progress[session_id] = f"[{session_id}] Обрабатываем артикулы..."
        print(f"[{session_id}] Начинаем обработку {len(df)} артикулов...", flush=True)
        results = [None] * len(df)
        
        with ThreadPoolExecutor(max_workers=3) as executor:
            futures = {}
            threading.current_thread().name = session_id
            for idx, article in enumerate(df.iloc[:, article_col]):
                futures[executor.submit(process_article, article, slot_names)] = idx
            processed = 0
            for fut in as_completed(futures):
                idx = futures[fut]
                if cancel_flags.get(session_id):
                    progress[session_id] = f"[{session_id}] Процесс отменён пользователем"
                    return
                results[idx] = fut.result()
                processed += 1
                if processed % 5 == 0 or processed == len(df):  # Обновляем чаще
                    progress[session_id] = f"[{session_id}] Обработано {processed}/{len(df)}"
                    print(f"[{session_id}] Обработано артикулов: {processed}/{len(df)}", flush=True)

        # Формируем итоговую таблицу
        progress[session_id] = f"[{session_id}] Формируем итоговую таблицу..."
        print(f"[{session_id}] Формируем итоговую таблицу...", flush=True)
        data = []
        for i, (art, name, slots_text) in enumerate(results):
            if cancel_flags.get(session_id):
                progress[session_id] = f"[{session_id}] Процесс отменён пользователем"
                return
            sticker = str(df.iat[i, sticker_col]).strip() if pd.notna(df.iat[i, sticker_col]) else ''
            qty = df.iat[i, quantity_col] if pd.notna(df.iat[i, quantity_col]) else 0
            data.append({
                '№ Стикера': sticker,
                'Количество': qty,
                'Артикул': art or '',
                'Ячейки склада': slots_text,
                'Название': name
            })

        # Сохраняем DataFrame
        progress[session_id] = f"[{session_id}] Сохраняем Excel файл..."
        print(f"[{session_id}] Создаем DataFrame и сохраняем в Excel...", flush=True)
        out_df = pd.DataFrame(data)[['№ Стикера','Количество','Артикул','Ячейки склада','Название']]
        out_df.to_excel(output_path, index=False)
        progress[session_id] = f"[{session_id}] Файл сохранён, начинаем форматирование..."
        print(f"[{session_id}] Excel файл сохранен, начинаем форматирование...", flush=True)

        if cancel_flags.get(session_id):
            progress[session_id] = f"[{session_id}] Процесс отменён пользователем"
            return

        # Форматирование
        progress[session_id] = f"[{session_id}] Загружаем файл для форматирования..."
        print(f"[{session_id}] Загружаем workbook для форматирования...", flush=True)
        
        try:
            wb = load_workbook(output_path)
            ws = wb.active
            progress[session_id] = f"[{session_id}] Workbook загружен, применяем границы..."
            print(f"[{session_id}] Workbook загружен, размеры: {ws.max_row}x{ws.max_column}", flush=True)
            
            # Применяем границы
            thin = Side(border_style='thin', color='000000')
            border = Border(left=thin, right=thin, top=thin, bottom=thin)
            max_row, max_col = ws.max_row, ws.max_column
            
            print(f"[{session_id}] Применяем границы к {max_row * max_col} ячейкам...", flush=True)
            progress[session_id] = f"[{session_id}] Применяем границы к {max_row * max_col} ячейкам..."
            
            cells_processed = 0
            for row in ws.iter_rows(min_row=1, max_row=max_row, min_col=1, max_col=max_col):
                if cancel_flags.get(session_id):
                    progress[session_id] = f"[{session_id}] Процесс отменён пользователем"
                    return
                for cell in row:
                    cell.border = border
                    cells_processed += 1
                    if cells_processed % 100 == 0:  # Обновляем прогресс каждые 100 ячеек
                        progress[session_id] = f"[{session_id}] Обработано границ: {cells_processed}/{max_row * max_col}"

            # Форматируем стикеры
            progress[session_id] = f"[{session_id}] Форматируем номера стикеров..."
            print(f"[{session_id}] Форматируем {max_row-1} номеров стикеров...", flush=True)
            
            for r in range(2, max_row+1):
                if cancel_flags.get(session_id):
                    progress[session_id] = f"[{session_id}] Процесс отменён пользователем"
                    return
                try:
                    format_sticker_cell(ws.cell(row=r, column=1))
                    if r % 50 == 0:  # Обновляем прогресс каждые 50 строк
                        progress[session_id] = f"[{session_id}] Отформатировано стикеров: {r-1}/{max_row-1}"
                except Exception as e:
                    print(f"[{session_id}] Ошибка форматирования строки {r}: {e}", flush=True)

            # Сохраняем с повторными попытками
            progress[session_id] = f"[{session_id}] Сохраняем отформатированный файл..."
            print(f"[{session_id}] Сохраняем отформатированный файл...", flush=True)
            
            if not save_workbook_with_retries(wb, output_path, session_id):
                progress[session_id] = f"[{session_id}] Ошибка: не удалось сохранить файл"
                return

        except Exception as e:
            error_msg = f"Ошибка форматирования: {str(e)}"
            progress[session_id] = f"[{session_id}] {error_msg}"
            print(f"[{session_id}] ОШИБКА ФОРМАТИРОВАНИЯ: {error_msg}", flush=True)
            # Продолжаем без форматирования, файл уже сохранен

        if cancel_flags.get(session_id):
            progress[session_id] = f"[{session_id}] Процесс отменён пользователем"
            return

        # Очищаем папку результатов
        progress[session_id] = f"[{session_id}] Очищаем старые файлы..."
        print(f"[{session_id}] Очищаем старые файлы...", flush=True)
        clean_old_results(max_files=50)
        
        progress[session_id] = f"[{session_id}] Обработка завершена успешно!"
        print(f"[{session_id}] ✅ Обработка завершена успешно!", flush=True)

    except Exception as e:
        error_msg = f"Критическая ошибка обработки: {str(e)}"
        progress[session_id] = f"[{session_id}] {error_msg}"
        print(f"[{session_id}] КРИТИЧЕСКАЯ ОШИБКА: {error_msg}", flush=True)
        import traceback
        traceback.print_exc()

# =================== HTTP Routes ===================

@app.route('/', methods=['GET','POST'])
def upload_file():
    if request.method == 'POST':
        if 'file' not in request.files:
            flash('Файл не выбран')
            return redirect(request.url)
        file = request.files['file']
        if not file or file.filename == '':
            flash('Файл не выбран')
            return redirect(request.url)
        filename = file.filename
        session_id = str(time.time())
        inp = os.path.join(UPLOAD_FOLDER, session_id+"_"+filename)
        out = os.path.join(RESULT_FOLDER, f"result_{session_id}.xlsx")
        file.save(inp)
        t = threading.Thread(target=process_file, args=(inp, out, session_id), name=session_id)
        t.start()
        return render_template_string(HEADER_HTML + '''
<script>sessionStorage.setItem('currentSession',''' + f"'{session_id}'" + ''');</script>
<meta http-equiv="refresh" content="0;url=/processing/''' + session_id + '''/result_''' + session_id + '''.xlsx">''')
    
    # Получаем список последних файлов для отображения
    recent_files = get_recent_files(10)
    files_html = ""
    if recent_files:
        files_html = '''
<h3 style="margin-top:30px;">Последние сгенерированные файлы:</h3>
<div style="border:1px solid #ddd; border-radius:5px; padding:15px; background:#f9f9f9;">
'''
        for i, file_info in enumerate(recent_files, 1):
            files_html += f'''
  <div style="display:flex; justify-content:space-between; align-items:center; padding:8px 0; border-bottom:1px solid #eee;">
    <div>
      <strong>{i}.</strong> 
      <span style="margin-left:10px;">{file_info['filename']}</span>
      <small style="margin-left:15px; color:#666;">({file_info['formatted_time']}, {file_info['formatted_size']})</small>
    </div>
    <a href="/download/{file_info['filename']}" style="background:#007bff; color:white; padding:5px 15px; text-decoration:none; border-radius:3px; font-size:12px;">Скачать</a>
  </div>
'''
        files_html += "</div>"
    
    return render_template_string(HEADER_HTML + '''
<!doctype html>
<title>Главная</title>
<h2>Загрузите Excel файл с данными</h2>
<form method="post" enctype="multipart/form-data">
  <input type="file" name="file" accept=".xlsx,.xls" required>
  <button type="submit">Загрузить</button>
</form>
''' + files_html)

@app.route('/create_order/<session_id>/<filename>', methods=['POST'])
def create_order(session_id, filename):
    """Создает заказ покупателя на основе данных из файла в отдельном потоке."""
    filepath = os.path.join(RESULT_FOLDER, filename)
    
    if not os.path.exists(filepath):
        return jsonify({"error": "Файл не найден"}), 404
    
    # Запускаем создание заказа в отдельном потоке
    order_session_id = f"order_{session_id}_{int(time.time())}"
    cancel_flags[f"order_{order_session_id}"] = False
    order_progress[order_session_id] = "🔄 Инициализация создания заказа..."
    
    def create_order_thread():
        try:
            result = create_customer_order_from_file(filepath, order_session_id)
            order_progress[order_session_id + "_result"] = result
        except Exception as e:
            order_progress[order_session_id + "_result"] = {"error": f"Критическая ошибка: {str(e)}"}
    
    thread = threading.Thread(target=create_order_thread, name=order_session_id)
    thread.start()
    
    return jsonify({"success": True, "order_session_id": order_session_id})

@app.route('/order_status/<order_session_id>')
def order_status(order_session_id):
    """Возвращает статус создания заказа."""
    status = order_progress.get(order_session_id, "Нет данных о создании заказа")
    result = order_progress.get(order_session_id + "_result")
    
    return jsonify({
        "status": status,
        "result": result,
        "completed": result is not None
    })

@app.route('/cancel_order/<order_session_id>', methods=['POST'])
def cancel_order(order_session_id):
    """Отменяет создание заказа."""
    cancel_flags[f"order_{order_session_id}"] = True
    order_progress[order_session_id] = "❌ Отмена создания заказа..."
    return jsonify({'status': 'cancelling'})

@app.route('/cancel/<session_id>', methods=['POST'])
def cancel(session_id):
    cancel_flags[session_id] = True
    progress[session_id] = f"[{session_id}] Процесс отменён пользователем"
    return jsonify({'status':'cancelled'})

@app.route('/status/<session_id>')
def status(session_id):
    return jsonify({'status': progress.get(session_id, 'Нет данных')})

@app.route('/processing/<session_id>/<filename>')
def processing(session_id, filename):
    return render_template_string(HEADER_HTML + f'''
<!doctype html>
<title>Обработка</title>
<h2>Статус обработки файла</h2>
<p id="status">Загрузка...</p>
<div style="margin-top:20px;">
  <button onclick="cancelProcess()" style="margin-right:10px;">Остановить процесс</button>
  <button id="createOrderBtn" onclick="createOrder()" disabled style="background-color:#ccc; cursor:not-allowed;">Создать заказ</button>
  <button id="cancelOrderBtn" onclick="cancelOrder()" disabled style="background-color:#dc3545; cursor:not-allowed; margin-left:10px; display:none;">Остановить создание заказа</button>
</div>
<div id="orderProgress" style="margin-top:15px; padding:10px; background:#f8f9fa; border-left:4px solid #007bff; display:none;"></div>
<div id="orderResult" style="margin-top:20px; display:none;"></div>
<script>
let orderSessionId = null;
let orderCheckInterval = null;

function checkStatus() {{
  fetch('/status/{session_id}')
    .then(r=>r.json()).then(data=>{{
       document.getElementById('status').innerText=data.status;
       const createOrderBtn = document.getElementById('createOrderBtn');
       if(data.status.toLowerCase().includes('завершена')) {{
         createOrderBtn.disabled = false;
         createOrderBtn.style.backgroundColor = '#007bff';
         createOrderBtn.style.cursor = 'pointer';
         createOrderBtn.style.color = 'white';
         window.location.href = '/download/{filename}';
       }} else if(data.status.toLowerCase().includes('ошибка') || data.status.toLowerCase().includes('отменён')) {{
         createOrderBtn.disabled = true;
         createOrderBtn.style.backgroundColor = '#dc3545';
         createOrderBtn.style.cursor = 'not-allowed';
         createOrderBtn.style.color = 'white';
         window.location.href = '/download/{filename}';
       }} else {{
         setTimeout(checkStatus,2000);
       }}
    }});
}}

function createOrder() {{
  const btn = document.getElementById('createOrderBtn');
  const cancelBtn = document.getElementById('cancelOrderBtn');
  const progressDiv = document.getElementById('orderProgress');
  
  if(!btn.disabled) {{
    btn.disabled = true;
    btn.style.backgroundColor = '#6c757d';
    btn.innerText = 'Создание...';
    
    cancelBtn.style.display = 'inline-block';
    cancelBtn.disabled = false;
    cancelBtn.style.cursor = 'pointer';
    
    progressDiv.style.display = 'block';
    progressDiv.innerHTML = '🔄 Инициализация создания заказа...';
    
    fetch('/create_order/{session_id}/{filename}', {{method:'POST'}})
      .then(r=>r.json())
      .then(data=>{{
        if(data.success) {{
          orderSessionId = data.order_session_id;
          orderCheckInterval = setInterval(checkOrderStatus, 1000);
        }} else {{
          showOrderError(data.error || 'Неизвестная ошибка запуска');
        }}
      }})
      .catch(error=>{{
        showOrderError('Ошибка соединения: ' + error.message);
      }});
  }}
}}

function checkOrderStatus() {{
  if(!orderSessionId) return;
  
  fetch(`/order_status/${{orderSessionId}}`)
    .then(r=>r.json())
    .then(data=>{{
      const progressDiv = document.getElementById('orderProgress');
      progressDiv.innerHTML = data.status;
      
      if(data.completed) {{
        clearInterval(orderCheckInterval);
        const cancelBtn = document.getElementById('cancelOrderBtn');
        cancelBtn.style.display = 'none';
        
        if(data.result && data.result.success) {{
          showOrderSuccess(data.result);
        }} else {{
          showOrderError(data.result ? data.result.error : 'Неизвестная ошибка');
        }}
      }}
    }})
    .catch(error=>{{
      console.error('Ошибка проверки статуса заказа:', error);
    }});
}}

function showOrderSuccess(result) {{
  const btn = document.getElementById('createOrderBtn');
  const resultDiv = document.getElementById('orderResult');
  const progressDiv = document.getElementById('orderProgress');
  
  progressDiv.style.display = 'none';
  resultDiv.innerHTML = `
    <div style="padding:15px; background:#d4edda; border:1px solid #c3e6cb; border-radius:5px; color:#155724;">
      <h4>✅ Заказ создан успешно!</h4>
      <p><strong>Название:</strong> ${{result.order_name}}</p>
      <p><strong>Позиций добавлено:</strong> ${{result.positions_added}} из ${{result.total_items}}</p>
      <p><strong>ID заказа:</strong> ${{result.order_id}}</p>
      ${{result.not_found_articles && result.not_found_articles.length > 0 ? 
        '<p><strong>⚠️ Не найдены артикулы:</strong> ' + result.not_found_articles.join(', ') + '</p>' : ''}}
    </div>
  `;
  btn.innerText = 'Заказ создан';
  btn.style.backgroundColor = '#28a745';
  resultDiv.style.display = 'block';
}}

function showOrderError(error) {{
  const btn = document.getElementById('createOrderBtn');
  const resultDiv = document.getElementById('orderResult');
  const progressDiv = document.getElementById('orderProgress');
  
  progressDiv.style.display = 'none';
  resultDiv.innerHTML = `
    <div style="padding:15px; background:#f8d7da; border:1px solid #f5c6cb; border-radius:5px; color:#721c24;">
      <h4>❌ Ошибка создания заказа</h4>
      <p>${{error}}</p>
    </div>
  `;
  btn.disabled = false;
  btn.style.backgroundColor = '#007bff';
  btn.innerText = 'Создать заказ';
  resultDiv.style.display = 'block';
}}

function cancelProcess() {{
  fetch('/cancel/{session_id}',{{method:'POST'}})
    .then(()=>alert('Процесс будет остановлен'));
}}

function cancelOrder() {{
  if(orderSessionId) {{
    fetch(`/cancel_order/${{orderSessionId}}`, {{method:'POST'}})
      .then(()=>{{
        clearInterval(orderCheckInterval);
        const progressDiv = document.getElementById('orderProgress');
        progressDiv.innerHTML = '❌ Отмена создания заказа...';
        
        setTimeout(() => {{
          const btn = document.getElementById('createOrderBtn');
          const cancelBtn = document.getElementById('cancelOrderBtn');
          btn.disabled = false;
          btn.style.backgroundColor = '#007bff';
          btn.innerText = 'Создать заказ';
          cancelBtn.style.display = 'none';
          progressDiv.style.display = 'none';
        }}, 2000);
      }});
  }}
}}

checkStatus();
</script>
''')

@app.route('/download/<filename>')
def download(filename):
    path = os.path.join(RESULT_FOLDER, filename)
    if os.path.exists(path):
        return send_file(path, as_attachment=True)
    flash('Файл не найден')
    return redirect('/')

if __name__ == '__main__':
    app.run(host='0.0.0.0', port=5001, debug=True)