import pandas as pd
import time
from concurrent.futures import ThreadPoolExecutor, as_completed
from openpyxl import load_workbook
from openpyxl.styles import Border, Side
from moysklad_api import get_product_uuid, get_store_slots, get_stock_by_slot, STORE_ID
from utils import (
    find_column_index,
    find_quantity_column,
    save_workbook_with_retries,
    format_sticker_cell
)

progress: dict[str, str] = {}

def process_file(input_path: str, output_path: str, session_id: str):
    progress[session_id] = "🔄 Старт обработки"
    df = pd.read_excel(input_path, dtype=str)
    total = len(df)
    progress[session_id] = f"📥 Загружено строк: {total}"

    # Найти нужные колонки
    art_col     = find_column_index(df.columns, ['артикул'])
    sticker_col = find_column_index(df.columns, ['№ стикера', 'номер стикера', 'стикер', 'номер'])
    qty_col     = find_quantity_column(df)
    if art_col is None or sticker_col is None or qty_col is None:
        progress[session_id] = "❌ Ошибка: не найдены колонки Артикул/№ Стикера/Кол-во"
        return

    progress[session_id] = "📦 Получаем ячейки склада..."
    slots = get_store_slots(STORE_ID)
    progress[session_id] = f"✅ Ячеек: {len(slots)}"

    # Параллельно обрабатываем артикули
    results = [("", "", "")] * total
    with ThreadPoolExecutor(max_workers=5) as executor:
        futures = {
            executor.submit(_process_row, df.iat[i, art_col], slots): i
            for i in range(total)
        }
        done = 0
        for fut in as_completed(futures):
            idx = futures[fut]
            try:
                results[idx] = fut.result()
            except Exception:
                results[idx] = ("", "", "")
            done += 1
            progress[session_id] = f"🔄 Обработано {done}/{total}"

    # Собираем итоговую таблицу
    out_data = []
    for i, (article, name, slots_text) in enumerate(results):
        sticker = df.iat[i, sticker_col] or ""
        quantity = df.iat[i, qty_col] or ""
        out_data.append({
            '№ Стикера':    sticker,
            'Количество':   quantity,
            'Артикул':      article or "",
            'Ячейки склада': slots_text,
            'Название':     name or ""
        })

    out_df = pd.DataFrame(out_data, columns=[
        '№ Стикера', 'Количество', 'Артикул', 'Ячейки склада', 'Название'
    ])
    out_df.to_excel(output_path, index=False)

    # Форматирование рамок и столбца стикеров
    wb = load_workbook(output_path)
    ws = wb.active
    thin = Side(border_style="thin", color="000000")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for row in ws.iter_rows(min_row=1,
                            max_row=ws.max_row,
                            min_col=1,
                            max_col=ws.max_column):
        for cell in row:
            cell.border = border

    for r in range(2, ws.max_row + 1):
        format_sticker_cell(ws.cell(row=r, column=1))

    save_workbook_with_retries(wb, output_path)
    progress[session_id] = "✅ Обработка завершена"

def _process_row(article: str, slots: dict[str, str]) -> tuple[str, str, str]:
    art = str(article).strip()
    if not art:
        return "", "", ""
    uuid, name = get_product_uuid(art)
    if not uuid:
        return art, "", ""
    stock_rows = get_stock_by_slot(uuid, STORE_ID)
    parts = []
    for r in stock_rows:
        slot_id = r.get('slotId')
        qty = r.get('stock', 0)
        if slot_id and qty and qty != 0:
            parts.append(f"{slots.get(slot_id, slot_id)} - {int(qty)} шт")
    time.sleep(0.1)  # чтобы не получить 429 Too Many Requests
    return art, name or "", ", ".join(parts)
