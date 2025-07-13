import os
import time
import math
import threading
from concurrent.futures import ThreadPoolExecutor, as_completed
from flask import Flask, request, render_template_string, send_file, flash, redirect, url_for
import pandas as pd
import requests
from openpyxl import load_workbook
from openpyxl.styles import Border, Side, Font

app = Flask(__name__)
app.secret_key = 'your_secret_key'

API_TOKEN = "f9be4985f5e3488716c040ca52b8e04c7c0f9e0b".strip()
API_TOKEN = API_TOKEN.encode('ascii', errors='ignore').decode()

STORE_ID = "241ed919-a631-11ee-0a80-07a9000bb947"

HEADERS = {
    "Authorization": f"Bearer {API_TOKEN}",
    "Accept-Encoding": "gzip",
    "Content-Type": "application/json"
}

UPLOAD_FOLDER = 'uploads'
RESULT_FOLDER = 'results'
os.makedirs(UPLOAD_FOLDER, exist_ok=True)
os.makedirs(RESULT_FOLDER, exist_ok=True)

progress = {}

def find_column_index(columns, names):
    for idx, col in enumerate(columns):
        col_lower = str(col).strip().lower()
        for name in names:
            if col_lower == name.lower():
                return idx
    return None

def find_quantity_column(df, candidates=['кол-во', 'количество', 'кол']):
    for col in df.columns[:10]:
        col_lower = str(col).strip().lower()
        if any(name == col_lower for name in candidates):
            sample = df[col].dropna().head(10)
            if not sample.empty and pd.to_numeric(sample, errors='coerce').notnull().any():
                return df.columns.get_loc(col)
    return None

def get_product_uuid(article):
    url = "https://api.moysklad.ru/api/remap/1.2/entity/product"
    params = {"filter": f"article={article}", "limit": 1}
    resp = requests.get(url, headers=HEADERS, params=params)
    resp.raise_for_status()
    data = resp.json()
    rows = data.get('rows', [])
    if not rows:
        return None, None
    row = rows[0]
    return row['id'], row.get('name', '')

def get_store_slots(store_id):
    url = f"https://api.moysklad.ru/api/remap/1.2/entity/store/{store_id}/slots"
    params = {"limit": 1000}
    resp = requests.get(url, headers=HEADERS, params=params)
    resp.raise_for_status()
    data = resp.json()
    return {row['id']: row['name'] for row in data.get('rows', [])}

def get_stock_by_slot(product_uuid, store_id):
    url = "https://api.moysklad.ru/api/remap/1.2/report/stock/byslot/current"
    params = [
        ('filter', f"assortmentId={product_uuid}"),
        ('filter', f"storeId={store_id}"),
        ('limit', '1000')
    ]
    resp = requests.get(url, headers=HEADERS, params=params)
    resp.raise_for_status()
    return resp.json()

def process_article(article, slot_names):
    article = str(article).strip()
    if not article:
        return None, None, ""

    try:
        uuid, name = get_product_uuid(article)
        if not uuid:
            return None, None, ""

        slot_rows = get_stock_by_slot(uuid, STORE_ID)
        parts = []
        for sr in slot_rows:
            slot_id = sr.get('slotId')
            qty = sr.get('stock', 0)
            if slot_id and qty > 0:
                parts.append(f"{slot_names.get(slot_id, slot_id)} - {int(qty)} шт")
        time.sleep(0.1)
        return article, name, ", ".join(parts)
    except Exception as e:
        print(f"Ошибка для артикула {article}: {e}", flush=True)
        return None, "", ""

def format_sticker_cell(cell):
    value = str(cell.value) if cell.value else ""
    if len(value) < 4:
        return
    main_part = value[:-4].rstrip()
    last_four = value[-4:]
    cell.value = f"{main_part} {last_four}"
    cell.font = cell.font.copy(bold=True, size=cell.font.size + 1)

def save_workbook_with_retries(wb, filename, retries=5, delay=3):
    for attempt in range(1, retries + 1):
        try:
            wb.save(filename)
            print(f"Файл {filename} успешно сохранён.", flush=True)
            return True
        except PermissionError:
            print(f"Попытка {attempt}: Файл {filename} занят или открыт. Жду {delay} сек...", flush=True)
            time.sleep(delay)
        except Exception as e:
            print(f"Ошибка сохранения файла: {e}", flush=True)
            return False
    print(f"Не удалось сохранить файл {filename} после {retries} попыток.", flush=True)
    return False

def process_file(input_path, output_path, session_id):
    progress[session_id] = f"[{session_id}] Начинаем обработку файла"
    df = pd.read_excel(input_path)
    progress[session_id] = f"[{session_id}] Загружен Excel, строк: {len(df)}"

    article_col = find_column_index(df.columns, ['артикул'])
    sticker_col = find_column_index(df.columns, ['№ стикера', 'номер стикера', 'стикер', 'номер'])
    quantity_col = find_quantity_column(df)
    if article_col is None or sticker_col is None or quantity_col is None:
        progress[session_id] = f"[{session_id}] Ошибка: не найдены обязательные колонки"
        return

    progress[session_id] = f"[{session_id}] Получаем список ячеек адресного склада"
    slot_names = get_store_slots(STORE_ID)
    progress[session_id] = f"[{session_id}] Получено ячеек: {len(slot_names)}"

    progress[session_id] = f"[{session_id}] Начинаем обработку артикулов..."
    with ThreadPoolExecutor(max_workers=3) as executor:
        futures = {
            executor.submit(process_article, article, slot_names): idx
            for idx, article in enumerate(df.iloc[:, article_col])
        }
        results = [("", "", "")] * len(df)
        processed_count = 0
        for future in as_completed(futures):
            idx = futures[future]
            try:
                results[idx] = future.result()
            except Exception as e:
                print(f"[{session_id}] Ошибка обработки артикула на строке {idx + 2}: {e}", flush=True)
                results[idx] = ("", "", "")
            processed_count += 1
            if processed_count % 10 == 0 or processed_count == len(df):
                progress[session_id] = f"[{session_id}] Обработано {processed_count} из {len(df)} артикулов"

    progress[session_id] = f"[{session_id}] Формируем итоговую таблицу..."
    data_for_df = []
    for i, (article, name, slots_text) in enumerate(results):
        sticker_val = ""
        if sticker_col is not None:
            val = df.iat[i, sticker_col]
            if pd.notna(val):
                sticker_val = str(val).strip()

        quantity_val = df.iat[i, quantity_col] if quantity_col is not None else 0

        data_for_df.append({
            '№ Стикера': sticker_val,
            'Количество': quantity_val,
            'Артикул': article if article else "",
            'Ячейки склада': slots_text,
            'Название': name
        })

    out_df = pd.DataFrame(data_for_df)
    progress[session_id] = f"[{session_id}] Записываем результат в Excel..."
    out_df = out_df[['№ Стикера', 'Количество', 'Артикул', 'Ячейки склада', 'Название']]
    out_df.to_excel(output_path, index=False)

    progress[session_id] = f"[{session_id}] Загружаем файл для форматирования..."
    wb = load_workbook(output_path)
    ws = wb.active

    thin = Side(border_style="thin", color="000000")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    max_row = ws.max_row
    max_col = ws.max_column
    progress[session_id] = f"[{session_id}] Форматируем обводку для {max_row} строк и {max_col} столбцов..."

    for row in ws.iter_rows(min_row=1, max_row=max_row, min_col=1, max_col=max_col):
        for cell in row:
            cell.border = border

    progress[session_id] = f"[{session_id}] Форматируем колонку '№ Стикера' ..."
    for row_idx in range(2, max_row + 1):
        cell = ws.cell(row=row_idx, column=1)
        format_sticker_cell(cell)

    progress[session_id] = f"[{session_id}] Сохраняем итоговый файл..."
    save_workbook_with_retries(wb, output_path)
    progress[session_id] = f"[{session_id}] Обработка файла завершена!"

@app.route('/', methods=['GET', 'POST'])
def upload_file():
    if request.method == 'POST':
        if 'file' not in request.files:
            flash('Файл не выбран')
            return redirect(request.url)
        file = request.files['file']
        if file.filename == '':
            flash('Файл не выбран')
            return redirect(request.url)
        if file:
            filename = file.filename
            session_id = str(time.time())
            input_path = os.path.join(UPLOAD_FOLDER, session_id + "_" + filename)
            output_filename = f"result_{session_id}.xlsx"
            output_path = os.path.join(RESULT_FOLDER, output_filename)

            file.save(input_path)

            thread = threading.Thread(target=process_file, args=(input_path, output_path, session_id))
            thread.start()

            return redirect(url_for('processing', session_id=session_id, filename=output_filename))

    return render_template_string('''
    <!doctype html>
    <title>Загрузка файла для обработки</title>
    <h1>Загрузите Excel файл с данными</h1>
    {% with messages = get_flashed_messages() %}
      {% if messages %}
        <ul>
        {% for message in messages %}
          <li style="color:red;">{{ message }}</li>
        {% endfor %}
        </ul>
      {% endif %}
    {% endwith %}
    <form method=post enctype=multipart/form-data>
      <input type=file name=file accept=".xlsx,.xls" required>
      <input type=submit value=Загрузить>
    </form>
    ''')

@app.route('/status/<session_id>')
def status(session_id):
    status_msg = progress.get(session_id, "Нет данных о статусе")
    return {"status": status_msg}

@app.route('/processing/<session_id>/<filename>')
def processing(session_id, filename):
    output_path = os.path.join(RESULT_FOLDER, filename)
    return render_template_string(f'''
    <!doctype html>
    <title>Обработка файла</title>
    <h1>Статус обработки файла</h1>
    <p id="status">Загрузка...</p>
    <script>
        function checkStatus() {{
            fetch('/status/{session_id}')
            .then(response => response.json())
            .then(data => {{
                document.getElementById('status').innerText = data.status;
                if (data.status.toLowerCase().includes("завершена")) {{
                    window.location.href = "/download/{filename}";
                }} else {{
                    setTimeout(checkStatus, 2000);
                }}
            }});
        }}
        checkStatus();
    </script>
    ''')

@app.route('/download/<filename>')
def download(filename):
    path = os.path.join(RESULT_FOLDER, filename)
    if os.path.exists(path):
        return send_file(path, as_attachment=True)
    else:
        flash("Файл не найден.")
        return redirect('/')

if __name__ == '__main__':
    app.run(host='0.0.0.0', port=5001, debug=True)
